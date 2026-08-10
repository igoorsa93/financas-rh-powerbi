# -*- coding: utf-8 -*-
"""
Gera versoes ficticias simplificadas dos 12 arquivos-fonte originais
(10 tipos de fonte), usando os mesmos funcionarios/valores ficticios
usados no RH_Folha_Consolidado.xlsx.
"""
import json
import csv
import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill

random.seed(7)

BASE = "C:/Users/igoor/Desktop/Controle Pessoal/github-"
OUT_DIR = f"{BASE}/Arquivos Fonte RH"

with open(f"{BASE}/scripts/funcionarios_fake.json", encoding="utf-8") as f:
    FUNCS = json.load(f)

HEADER_FILL = PatternFill("solid", fgColor="1E5B34")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")


def novo_wb():
    wb = openpyxl.Workbook()
    ws = wb.active
    return wb, ws


def escreve_cabecalho(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def salvar(wb, nome):
    caminho = f"{OUT_DIR}/{nome}"
    wb.save(caminho)
    print("Salvo:", caminho)


def salvar_csv(nome, headers, linhas):
    caminho = f"{OUT_DIR}/{nome}"
    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(headers)
        w.writerows(linhas)
    print("Salvo:", caminho)


depts_ordered = list(dict.fromkeys(f["departamento"] for f in FUNCS))


# 1) Relacao de Empregados (era .xls) -> relacao_empregados.xlsx
def gerar_relacao_empregados():
    wb, ws = novo_wb()
    ws.title = "Empregados"
    escreve_cabecalho(ws, ["Codigo", "Nome", "Cargo", "Admissao", "Salario", "Servico", "Departamento", "C_de_Custo", "Nascimento", "Sexo"])
    for i, f in enumerate(FUNCS):
        r = i + 2
        ws.append([f["codigo"], f["nome"], f["cargo"], f["admissao"], f["salario"], f["servico"], f["departamento"], f["c_de_custo"], f["nascimento"], f["sexo"]])
    salvar(wb, "relacao_empregados.xlsx")


# 2) Relacao de Faltas (era .xls) -> relacao_de_faltas.xlsx
def gerar_relacao_faltas():
    wb, ws = novo_wb()
    ws.title = "Faltas"
    escreve_cabecalho(ws, ["Funcionario", "Previstas_h", "Trabalhadas_h", "Ausencias_h", "Pct_Ausencias"])
    for f in FUNCS:
        previstas = 220
        ausencias = round(random.uniform(0, 12), 2) if random.random() < 0.35 else 0
        ws.append([f["nome"], previstas, previstas - ausencias, ausencias, round(ausencias / previstas, 4)])
    salvar(wb, "relacao_de_faltas.xlsx")


# 3) Relacao de Atestados (era .xls) -> relacao_de_atestados.xlsx
def gerar_relacao_atestados():
    wb, ws = novo_wb()
    ws.title = "Atestados"
    escreve_cabecalho(ws, ["Funcionario", "Previstas_h", "Trabalhadas_h", "Ausencias_h", "Pct_Ausencias"])
    for f in FUNCS:
        previstas = 220
        ausencias = round(random.uniform(0, 10), 2) if random.random() < 0.25 else 0
        ws.append([f["nome"], previstas, previstas - ausencias, ausencias, round(ausencias / previstas, 4)])
    salvar(wb, "relacao_de_atestados.xlsx")


# 4-6) Relatorio de Liquidos (PDF originais, 3 variantes) -> CSV
def gerar_liquidos(nome_arquivo, tipo_folha, fator_min, fator_max, amostragem=1.0):
    headers = ["Departamento", "Codigo", "Nome", "Valor", "Data_Pagamento", "Tipo Folha"]
    linhas = []
    for f in FUNCS:
        if random.random() > amostragem:
            continue
        valor = round(f["salario"] * random.uniform(fator_min, fator_max), 2)
        linhas.append([f["departamento"], f["codigo"], f["nome"], valor, "01/06/2026", tipo_folha])
    salvar_csv(nome_arquivo, headers, linhas)


# 7-8) Relatorios de Hora Extra (PDF originais, 2 variantes) -> CSV
def gerar_horas_extras(nome_arquivo, tipo_folha, amostragem=0.6):
    headers = ["Departamento", "Codigo", "Nome", "Rubrica_Nome", "Competencia", "Valor_Calculado", "Horas"]
    linhas = []
    for f in random.sample(FUNCS, k=int(len(FUNCS) * amostragem)):
        horas = round(random.uniform(1, 20), 2)
        valor = round(horas * (f["salario"] / 220) * 1.5, 2)
        linhas.append([f["departamento"], f["codigo"], f["nome"], random.choice(["HORAS EXTRAS 50%", "HORAS EXTRAS 100%"]), "05/2026", valor, horas])
    salvar_csv(nome_arquivo, headers, linhas)


# 9) Relatorios de Falta por funcionario (PDF original) -> CSV
def gerar_faltas_por_funcionario():
    headers = ["Departamento", "Codigo", "Nome", "Rubrica_Nome", "Competencia", "Valor_Calculado", "Horas"]
    linhas = []
    for f in random.sample(FUNCS, k=8):
        horas = round(random.uniform(2, 16), 2)
        valor = round(horas * (f["salario"] / 220), 2)
        linhas.append([f["departamento"], f["codigo"], f["nome"], "HORAS FALTAS", "05/2026", valor, horas])
    salvar_csv("relatorios_de_falta_por_funcionario.csv", headers, linhas)


# 10) Relacao de Rescisoes Calculadas (PDF original) -> CSV
def gerar_rescisoes():
    headers = ["Codigo", "Empregado", "Admissao", "Aviso", "Demissao", "Saldo_FGTS", "Salario", "Proventos", "Descontos", "Liquido", "FGTS_Rescisorio", "Motivo_Demissao"]
    linhas = []
    motivos = ["Pedido de Demissao", "Dispensa sem justa causa", "Resc. cont. exp. antec. empregador", "Termino de Contrato"]
    for f in random.sample(FUNCS, k=6):
        demissao = date(2026, 5, random.randint(1, 28))
        aviso = demissao - timedelta(days=random.randint(0, 30))
        proventos = round(f["salario"] * random.uniform(0.45, 0.7), 2)
        descontos = round(proventos * random.uniform(0.2, 0.35), 2)
        linhas.append([f["codigo"], f["nome"], f["admissao"], aviso.isoformat(), demissao.isoformat(),
                        round(random.uniform(50, 300), 2), f["salario"], proventos, descontos,
                        round(proventos - descontos, 2), round(f["salario"] * 0.08, 2), random.choice(motivos)])
    salvar_csv("relacao_de_rescisoes_calculadas.csv", headers, linhas)


# 11-12) Resumo Mensal GERAL (PDF originais, Normal + Extra) -> CSV
def gerar_resumo_mensal(nome_arquivo, tipo_folha):
    headers = ["Departamento", "Rubrica_Nome", "Tipo", "N_Empregados", "Valor_Informado", "Valor_Calculado"]
    linhas = []
    for dept in depts_ordered:
        membros = [f for f in FUNCS if f["departamento"] == dept]
        n_emp = len(membros)
        if n_emp == 0:
            continue
        folha_base = sum(f["salario"] for f in membros)
        fator_folha = 1.0 if tipo_folha == "Normal" else 0.12
        for rnome, rtipo, fator in [
            ("HORAS NORMAIS", "Provento", 0.78), ("HORAS EXTRAS 50%", "Provento", 0.06),
            ("INSS", "Desconto", 0.09), ("IRRF", "Desconto", 0.04),
        ]:
            valor_informado = round(folha_base * fator_folha * fator, 2)
            valor_calc = round(valor_informado * random.uniform(0.95, 1.05), 2)
            linhas.append([dept, rnome, rtipo, n_emp, valor_informado, valor_calc])
    salvar_csv(nome_arquivo, headers, linhas)


if __name__ == "__main__":
    gerar_relacao_empregados()
    gerar_relacao_faltas()
    gerar_relacao_atestados()
    gerar_liquidos("relatorio_de_liquidos_folha_extra.csv", "Extra", 0.02, 0.08, amostragem=0.4)
    gerar_liquidos("relatorio_de_liquidos_folha_normal_principal.csv", "Normal Principal", 0.55, 0.62)
    gerar_liquidos("relatorio_de_liquidos_folha_normal_quinzena.csv", "Normal Quinzena", 0.38, 0.45)
    gerar_horas_extras("relatorios_de_hora_extra_folha_extra.csv", "Extra", amostragem=0.3)
    gerar_horas_extras("relatorios_de_hora_extra_folha_normal.csv", "Normal", amostragem=0.6)
    gerar_faltas_por_funcionario()
    gerar_rescisoes()
    gerar_resumo_mensal("resumo_mensal_geral_folha_normal.csv", "Normal")
    gerar_resumo_mensal("resumo_mensal_geral_folha_extra.csv", "Extra")
    print("Concluido.")
