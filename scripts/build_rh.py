# -*- coding: utf-8 -*-
"""
Constroi RH_Folha_Consolidado.xlsx com a mesma estrutura de abas e
formulas do workbook real, populado 100% com dados ficticios.
"""
import json
import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(7)

BASE = "C:/Users/igoor/Desktop/Controle Pessoal/github-"
with open(f"{BASE}/scripts/funcionarios_fake.json", encoding="utf-8") as f:
    FUNCS = json.load(f)

HEADER_FILL = PatternFill("solid", fgColor="1E5B34")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=20, bold=True, color="153E24")
NORMAL_FONT = Font(name="Arial", size=10)

RUBRICAS_HORAS_EXTRAS = ["HORAS EXTRAS 50%", "HORAS EXTRAS 100%", "HORAS EXTRAS NOTURNAS"]
RUBRICAS_FALTAS = ["HORAS FALTAS", "FALTAS INJUSTIFICADAS"]
RUBRICAS_FOLHA = [
    (1, "HORAS NORMAIS", "Provento"),
    (2, "HORAS EXTRAS 50%", "Provento"),
    (3, "ADICIONAL NOTURNO", "Provento"),
    (4, "DSR SOBRE HORAS EXTRAS", "Provento"),
    (5, "INSS", "Desconto"),
    (6, "IRRF", "Desconto"),
    (7, "VALE TRANSPORTE", "Desconto"),
    (8, "VALE ALIMENTACAO", "Desconto"),
]
MOTIVOS_DEMISSAO = [
    "Pedido de Demissao",
    "Dispensa sem justa causa",
    "Resc. cont. exp. antec. empregador",
    "Termino de Contrato",
]


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---------------- Capa ----------------
    ws = wb.create_sheet("Capa")
    ws["B5"] = "ACME PRE-MOLDADOS INTELIGENTES LTDA (DADOS FICTICIOS)"
    ws["B5"].font = TITLE_FONT
    ws["B6"] = "Controle de Folha de Pagamento - Consolidado RH ()"
    ws["B6"].font = Font(name="Arial", size=13, bold=True)
    ws["B7"] = "Competencia: 05/2026 (fictício)"
    ws["B7"].font = Font(name="Arial", size=11)
    ws["B9"] = "Abas do workbook"
    ws["B9"].font = Font(name="Arial", size=12, bold=True)
    abas_desc = [
        ("Ficha_Funcionario", "Consulta individual: selecione o codigo e veja tudo vinculado"),
        ("Funcionarios", "Cadastro completo (fonte de todos os vinculos)"),
        ("Liquidos", "Valores liquidos pagos por funcionario (Normal Principal/Quinzena/Extra)"),
        ("Horas_Extras", "Horas extras trabalhadas e valores"),
        ("Faltas / Atestados", "Absenteismo por funcionario"),
        ("Faltas_Detalhado", "Movimentos de falta por rubrica"),
        ("Rescisoes", "Rescisoes calculadas no periodo"),
        ("Rubricas_Folha", "Proventos/descontos por departamento e rubrica"),
        ("Resumo_Departamento", "Totais por departamento"),
        ("Resumo_Geral", "KPIs consolidados (formulas)"),
    ]
    for i, (aba, desc) in enumerate(abas_desc):
        r = 10 + i
        ws.cell(row=r, column=2, value=aba).font = Font(name="Arial", bold=True)
        ws.cell(row=r, column=4, value=desc).font = NORMAL_FONT
    ws.cell(row=22, column=2, value="AVISO: todos os dados neste arquivo sao ficticios, gerados para fins de demonstracao/portfolio.").font = Font(name="Arial", italic=True, size=9, color="808080")
    autosize(ws, {"A": 3, "B": 32, "C": 3, "D": 60})

    # ---------------- Funcionarios ----------------
    ws = wb.create_sheet("Funcionarios")
    ws["A1"] = "ACME - Cadastro de Funcionarios ()"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    headers = ["Codigo", "Nome", "Cargo", "Admissao", "Salario", "Servico", "Departamento", "C_de_Custo", "Nascimento", "Sexo"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))
    for i, fdata in enumerate(FUNCS):
        r = 3 + i
        ws.cell(row=r, column=1, value=fdata["codigo"])
        ws.cell(row=r, column=2, value=fdata["nome"])
        ws.cell(row=r, column=3, value=fdata["cargo"])
        ws.cell(row=r, column=4, value=fdata["admissao"])
        ws.cell(row=r, column=5, value=fdata["salario"])
        ws.cell(row=r, column=6, value=fdata["servico"])
        ws.cell(row=r, column=7, value=fdata["departamento"])
        ws.cell(row=r, column=8, value=fdata["c_de_custo"])
        ws.cell(row=r, column=9, value=fdata["nascimento"])
        ws.cell(row=r, column=10, value=fdata["sexo"])
    last_func_row = 2 + len(FUNCS)
    autosize(ws, {"A": 8, "B": 34, "C": 24, "D": 12, "E": 10, "F": 8, "G": 26, "H": 18, "I": 12, "J": 10})

    # ---------------- Liquidos ----------------
    ws = wb.create_sheet("Liquidos")
    ws["A1"] = "ACME - Valores Liquidos Pagos ()"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    headers = ["Departamento_Num", "Departamento", "Codigo", "Nome", "Valor", "Data_Pagamento", "Tipo Folha", "Cargo (vinculo)", "C. de Custo (vinculo)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))
    row = 3
    for tipo_folha in ["Normal Principal", "Normal Quinzena", "Extra"]:
        for fdata in FUNCS:
            if tipo_folha == "Extra" and random.random() < 0.6:
                continue  # nem todos tem folha extra
            base_salario = fdata["salario"]
            if tipo_folha == "Normal Principal":
                valor = round(base_salario * random.uniform(0.55, 0.62), 2)
            elif tipo_folha == "Normal Quinzena":
                valor = round(base_salario * random.uniform(0.38, 0.45), 2)
            else:
                valor = round(random.uniform(80, 450), 2)
            ws.cell(row=row, column=1, value=list(dict.fromkeys(f["departamento"] for f in FUNCS)).index(fdata["departamento"]) + 1)
            ws.cell(row=row, column=2, value=fdata["departamento"])
            ws.cell(row=row, column=3, value=fdata["codigo"])
            ws.cell(row=row, column=4, value=fdata["nome"])
            ws.cell(row=row, column=5, value=valor)
            ws.cell(row=row, column=6, value="2026-06-01")
            ws.cell(row=row, column=7, value=tipo_folha)
            ws.cell(row=row, column=8, value=f"=IFERROR(INDEX(Funcionarios!$C$3:$C${last_func_row},MATCH($C{row},Funcionarios!$A$3:$A${last_func_row},0)),\"\")")
            ws.cell(row=row, column=9, value=f"=IFERROR(INDEX(Funcionarios!$H$3:$H${last_func_row},MATCH($C{row},Funcionarios!$A$3:$A${last_func_row},0)),\"\")")
            row += 1
    last_liquidos_row = row - 1
    autosize(ws, {"A": 16, "B": 26, "C": 8, "D": 34, "E": 12, "F": 16, "G": 16, "H": 22, "I": 20})

    # ---------------- Horas_Extras ----------------
    ws = wb.create_sheet("Horas_Extras")
    ws["A1"] = "ACME - Horas Extras ()"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    headers = ["Departamento_Num", "Departamento", "Codigo", "Nome", "Rubrica_Num", "Rubrica_Nome", "Competencia",
               "Valor_Calculado", "Horas", "Tipo_Movimento", "Unidade", "Tipo Folha", "Cargo (vinculo)", "C. de Custo (vinculo)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))
    row = 3
    depts_ordered = list(dict.fromkeys(f["departamento"] for f in FUNCS))
    amostra_he = random.sample(FUNCS, k=min(28, len(FUNCS)))
    for fdata in amostra_he:
        for tipo_folha in ["Normal", "Extra"]:
            if random.random() < 0.5:
                continue
            horas = round(random.uniform(1, 20), 2)
            valor = round(horas * (fdata["salario"] / 220) * 1.5, 2)
            ws.cell(row=row, column=1, value=depts_ordered.index(fdata["departamento"]) + 1)
            ws.cell(row=row, column=2, value=fdata["departamento"])
            ws.cell(row=row, column=3, value=fdata["codigo"])
            ws.cell(row=row, column=4, value=fdata["nome"])
            ws.cell(row=row, column=5, value=150)
            ws.cell(row=row, column=6, value=random.choice(RUBRICAS_HORAS_EXTRAS))
            ws.cell(row=row, column=7, value="05/2026")
            ws.cell(row=row, column=8, value=valor)
            ws.cell(row=row, column=9, value=horas)
            ws.cell(row=row, column=10, value="P")
            ws.cell(row=row, column=11, value="Horas")
            ws.cell(row=row, column=12, value=tipo_folha)
            ws.cell(row=row, column=13, value=f"=IFERROR(INDEX(Funcionarios!$C$3:$C${last_func_row},MATCH($C{row},Funcionarios!$A$3:$A${last_func_row},0)),\"\")")
            ws.cell(row=row, column=14, value=f"=IFERROR(INDEX(Funcionarios!$H$3:$H${last_func_row},MATCH($C{row},Funcionarios!$A$3:$A${last_func_row},0)),\"\")")
            row += 1
    last_he_row = row - 1
    autosize(ws, {"A": 16, "B": 26, "C": 8, "D": 34, "E": 10, "F": 20, "G": 12, "H": 14, "I": 8, "J": 12, "K": 8, "L": 10, "M": 22, "N": 20})

    # ---------------- Faltas / Atestados ----------------
    def build_faltas_sheet(nome_aba, tipo_label):
        ws = wb.create_sheet(nome_aba)
        ws["A1"] = f"ACME - {tipo_label} ()"
        ws["A1"].font = Font(name="Arial", bold=True, size=12)
        headers = ["Funcionario", "Previstas_h", "Trabalhadas_h", "Ausencias_h", "Pct_Ausencias", "Tipo",
                   "Codigo (vinculo)", "Departamento (vinculo)", "Cargo (vinculo)"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=2, column=c, value=h)
        style_header(ws, 2, len(headers))
        for i, fdata in enumerate(FUNCS):
            r = 3 + i
            previstas = 220
            ausencias = round(random.uniform(0, 12), 2) if random.random() < 0.35 else 0
            trabalhadas = previstas - ausencias
            pct = round(ausencias / previstas, 4) if previstas else 0
            ws.cell(row=r, column=1, value=fdata["nome"])
            ws.cell(row=r, column=2, value=previstas)
            ws.cell(row=r, column=3, value=trabalhadas)
            ws.cell(row=r, column=4, value=ausencias)
            ws.cell(row=r, column=5, value=pct)
            ws.cell(row=r, column=6, value=tipo_label)
            ws.cell(row=r, column=7, value=f"=IFERROR(INDEX(Funcionarios!$A$3:$A${last_func_row},MATCH($A{r},Funcionarios!$B$3:$B${last_func_row},0)),\"\")")
            ws.cell(row=r, column=8, value=f"=IFERROR(INDEX(Funcionarios!$G$3:$G${last_func_row},MATCH($A{r},Funcionarios!$B$3:$B${last_func_row},0)),\"\")")
            ws.cell(row=r, column=9, value=f"=IFERROR(INDEX(Funcionarios!$C$3:$C${last_func_row},MATCH($A{r},Funcionarios!$B$3:$B${last_func_row},0)),\"\")")
        autosize(ws, {"A": 34, "B": 12, "C": 14, "D": 12, "E": 14, "F": 12, "G": 16, "H": 22, "I": 20})
        return 2 + len(FUNCS)

    last_faltas_row = build_faltas_sheet("Faltas", "Faltas")
    last_atestados_row = build_faltas_sheet("Atestados", "Atestados")

    # ---------------- Faltas_Detalhado ----------------
    ws = wb.create_sheet("Faltas_Detalhado")
    ws["A1"] = "ACME - Faltas Detalhado ()"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    headers = ["Departamento_Num", "Departamento", "Codigo", "Nome", "Rubrica_Num", "Rubrica_Nome",
               "Competencia", "Valor_Calculado", "Horas", "Tipo_Movimento", "Unidade"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))
    amostra_falta = random.sample(FUNCS, k=8)
    for i, fdata in enumerate(amostra_falta):
        r = 3 + i
        horas = round(random.uniform(2, 16), 2)
        valor = round(horas * (fdata["salario"] / 220), 2)
        ws.cell(row=r, column=1, value=depts_ordered.index(fdata["departamento"]) + 1)
        ws.cell(row=r, column=2, value=fdata["departamento"])
        ws.cell(row=r, column=3, value=fdata["codigo"])
        ws.cell(row=r, column=4, value=fdata["nome"])
        ws.cell(row=r, column=5, value=40)
        ws.cell(row=r, column=6, value=random.choice(RUBRICAS_FALTAS))
        ws.cell(row=r, column=7, value="05/2026")
        ws.cell(row=r, column=8, value=valor)
        ws.cell(row=r, column=9, value=horas)
        ws.cell(row=r, column=10, value="D")
        ws.cell(row=r, column=11, value="Horas")
    autosize(ws, {"A": 16, "B": 26, "C": 8, "D": 34, "E": 10, "F": 18, "G": 12, "H": 14, "I": 8, "J": 12, "K": 8})

    # ---------------- Rescisoes ----------------
    ws = wb.create_sheet("Rescisoes")
    ws["A1"] = "ACME - Rescisoes ()"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    headers = ["Codigo", "Empregado", "Admissao", "Aviso", "Demissao", "Saldo_FGTS", "Salario",
               "Proventos", "Descontos", "Liquido", "FGTS_Rescisorio", "Motivo_Demissao",
               "Departamento (vinculo)", "Cargo (vinculo)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))
    amostra_resc = random.sample(FUNCS, k=6)
    for i, fdata in enumerate(amostra_resc):
        r = 3 + i
        demissao = date(2026, 5, random.randint(1, 28))
        aviso = demissao - timedelta(days=random.randint(0, 30))
        salario = fdata["salario"]
        proventos = round(salario * random.uniform(0.45, 0.7), 2)
        descontos = round(proventos * random.uniform(0.2, 0.35), 2)
        liquido = round(proventos - descontos, 2)
        ws.cell(row=r, column=1, value=fdata["codigo"])
        ws.cell(row=r, column=2, value=fdata["nome"])
        ws.cell(row=r, column=3, value=fdata["admissao"])
        ws.cell(row=r, column=4, value=aviso.isoformat())
        ws.cell(row=r, column=5, value=demissao.isoformat())
        ws.cell(row=r, column=6, value=round(random.uniform(50, 300), 2))
        ws.cell(row=r, column=7, value=salario)
        ws.cell(row=r, column=8, value=proventos)
        ws.cell(row=r, column=9, value=descontos)
        ws.cell(row=r, column=10, value=liquido)
        ws.cell(row=r, column=11, value=round(salario * 0.08, 2))
        ws.cell(row=r, column=12, value=random.choice(MOTIVOS_DEMISSAO))
        ws.cell(row=r, column=13, value=f"=IFERROR(INDEX(Funcionarios!$G$3:$G${last_func_row},MATCH($A{r},Funcionarios!$A$3:$A${last_func_row},0)),\"\")")
        ws.cell(row=r, column=14, value=f"=IFERROR(INDEX(Funcionarios!$C$3:$C${last_func_row},MATCH($A{r},Funcionarios!$A$3:$A${last_func_row},0)),\"\")")
    last_resc_row = 2 + len(amostra_resc)
    autosize(ws, {"A": 8, "B": 34, "C": 12, "D": 12, "E": 12, "F": 12, "G": 10, "H": 12, "I": 12, "J": 12, "K": 16, "L": 30, "M": 22, "N": 20})

    # ---------------- Rubricas_Folha ----------------
    ws = wb.create_sheet("Rubricas_Folha")
    ws["A1"] = "ACME - Rubricas da Folha ()"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    headers = ["Departamento_Num", "Departamento", "Rubrica_Num", "Rubrica_Nome", "Tipo",
               "N_Empregados", "Valor_Informado", "Valor_Calculado", "Tipo Folha"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))
    row = 3
    for dnum, dept in enumerate(depts_ordered, start=1):
        membros = [f for f in FUNCS if f["departamento"] == dept]
        n_emp = len(membros)
        if n_emp == 0:
            continue
        folha_base = sum(f["salario"] for f in membros)
        for tipo_folha in ["Normal", "Extra"]:
            fator_folha = 1.0 if tipo_folha == "Normal" else 0.12
            for rnum, rnome, rtipo in RUBRICAS_FOLHA:
                if tipo_folha == "Extra" and rtipo == "Provento" and rnum not in (2, 3):
                    continue
                if rtipo == "Provento":
                    # proventos maiores (base da folha), rubricas menores para adicionais
                    fator_rubrica = {1: 0.78, 2: 0.06, 3: 0.03, 4: 0.02}.get(rnum, 0.05)
                else:
                    # descontos como fracao menor do total de proventos da folha
                    fator_rubrica = {5: 0.09, 6: 0.04, 7: 0.02, 8: 0.015}.get(rnum, 0.02)
                valor_informado = round(folha_base * fator_folha * fator_rubrica, 2)
                valor_calc = round(valor_informado * random.uniform(0.95, 1.05), 2)
                ws.cell(row=row, column=1, value=dnum)
                ws.cell(row=row, column=2, value=dept)
                ws.cell(row=row, column=3, value=rnum)
                ws.cell(row=row, column=4, value=rnome)
                ws.cell(row=row, column=5, value=rtipo)
                ws.cell(row=row, column=6, value=n_emp)
                ws.cell(row=row, column=7, value=valor_informado)
                ws.cell(row=row, column=8, value=valor_calc)
                ws.cell(row=row, column=9, value=tipo_folha)
                row += 1
    last_rubricas_row = row - 1
    autosize(ws, {"A": 16, "B": 26, "C": 10, "D": 22, "E": 10, "F": 12, "G": 14, "H": 14, "I": 10})

    # ---------------- Resumo_Departamento ----------------
    ws = wb.create_sheet("Resumo_Departamento")
    ws["A1"] = "ACME - Resumo por Departamento ()"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    headers = ["Departamento_Num", "Departamento", "Tipo Folha", "Total_Proventos", "Total_Descontos", "Liquido_Depto"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))
    row = 3
    for dnum, dept in enumerate(depts_ordered, start=1):
        n_emp = sum(1 for f in FUNCS if f["departamento"] == dept)
        if n_emp == 0:
            continue
        for tipo_folha in ["Normal", "Extra"]:
            proventos = f"=SUMIFS(Rubricas_Folha!$H$3:$H${last_rubricas_row},Rubricas_Folha!$B$3:$B${last_rubricas_row},B{row},Rubricas_Folha!$E$3:$E${last_rubricas_row},\"Provento\",Rubricas_Folha!$I$3:$I${last_rubricas_row},C{row})"
            descontos = f"=SUMIFS(Rubricas_Folha!$H$3:$H${last_rubricas_row},Rubricas_Folha!$B$3:$B${last_rubricas_row},B{row},Rubricas_Folha!$E$3:$E${last_rubricas_row},\"Desconto\",Rubricas_Folha!$I$3:$I${last_rubricas_row},C{row})"
            ws.cell(row=row, column=1, value=dnum)
            ws.cell(row=row, column=2, value=dept)
            ws.cell(row=row, column=3, value=tipo_folha)
            ws.cell(row=row, column=4, value=proventos)
            ws.cell(row=row, column=5, value=descontos)
            ws.cell(row=row, column=6, value=f"=D{row}-E{row}")
            row += 1
    last_resumo_dep_row = row - 1
    autosize(ws, {"A": 16, "B": 26, "C": 12, "D": 16, "E": 16, "F": 14})

    # ---------------- Resumo_Geral ----------------
    ws = wb.create_sheet("Resumo_Geral")
    ws["A1"] = "Resumo Geral - Folha de Pagamento (Competencia 05/2026) ()"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws["A3"] = "Indicador"
    ws["B3"] = "Valor"
    ws["C3"] = "Observacao"
    style_header(ws, 3, 3)
    linhas = [
        ("Total Liquido Geral (Resumo_Departamento)", f"=SUM(Resumo_Departamento!F3:F{last_resumo_dep_row})", "Soma de Liquido_Depto (todas as folhas Normal+Extra), fonte: Resumo Mensal GERAL"),
        ("Total Liquido Geral (Liquidos - Normal Principal+Quinzena+Extra)", f"=SUM(Liquidos!E3:E{last_liquidos_row})", "Soma de Valor pago em Liquidos (3 relatorios de liquidos)"),
        ("Total de Funcionarios Ativos", f"=COUNTA(Funcionarios!A3:A{last_func_row})", "Contagem de linhas em Funcionarios (Relacao de Empregados)"),
        ("Total Horas Extras (horas)", f"=SUM(Horas_Extras!I3:I{last_he_row})", "Soma da coluna Horas em Horas_Extras (folhas Normal+Extra)"),
        ("Total Horas Extras (R$ calculado)", f"=SUM(Horas_Extras!H3:H{last_he_row})", "Soma da coluna Valor_Calculado em Horas_Extras"),
        ("Total Faltas (horas)", f"=SUM(Faltas!D3:D{last_faltas_row})", "Soma de Ausencias_h em Faltas (RELACAO DE FALTAS)"),
        ("Total Atestados (horas)", f"=SUM(Atestados!D3:D{last_atestados_row})", "Soma de Ausencias_h em Atestados (RELACAO DE ATESTADOS)"),
        ("Total Rescisoes (R$ Liquido)", f"=SUM(Rescisoes!J3:J{last_resc_row})", "Soma de Liquido em Rescisoes"),
    ]
    for i, (ind, val, obs) in enumerate(linhas):
        r = 4 + i
        ws.cell(row=r, column=1, value=ind)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=3, value=obs)
    ws.cell(row=13, column=1, value="Total Liquido por Departamento (Normal + Extra)").font = Font(bold=True)
    ws.cell(row=14, column=1, value="Departamento")
    ws.cell(row=14, column=2, value="Liquido Total")
    style_header(ws, 14, 2)
    for i, dept in enumerate(depts_ordered):
        r = 15 + i
        ws.cell(row=r, column=1, value=dept)
        ws.cell(row=r, column=2, value=f"=SUMIFS(Resumo_Departamento!$F$3:$F${last_resumo_dep_row},Resumo_Departamento!$B$3:$B${last_resumo_dep_row},A{r})")
    autosize(ws, {"A": 46, "B": 22, "C": 60})

    # ---------------- Ficha_Funcionario ----------------
    ws = wb.create_sheet("Ficha_Funcionario")
    ws["B2"] = "Ficha do Funcionario"
    ws["B2"].font = Font(name="Arial", bold=True, size=14)
    ws["B3"] = "Selecione o Codigo do funcionario na celula abaixo (dropdown) para ver todos os dados vinculados. ()"
    ws["B3"].font = Font(name="Arial", italic=True, size=9)
    ws["B5"] = "Codigo:"
    ws["C5"] = FUNCS[0]["codigo"]
    campos = [
        ("Nome", f"=IFERROR(INDEX(Funcionarios!$B$3:$B${last_func_row},MATCH($C$5,Funcionarios!$A$3:$A${last_func_row},0)),\"\")"),
        ("Cargo", f"=IFERROR(INDEX(Funcionarios!$C$3:$C${last_func_row},MATCH($C$5,Funcionarios!$A$3:$A${last_func_row},0)),\"\")"),
        ("Admissao", f"=IFERROR(INDEX(Funcionarios!$D$3:$D${last_func_row},MATCH($C$5,Funcionarios!$A$3:$A${last_func_row},0)),\"\")"),
        ("Salario", f"=IFERROR(INDEX(Funcionarios!$E$3:$E${last_func_row},MATCH($C$5,Funcionarios!$A$3:$A${last_func_row},0)),\"\")"),
        ("Departamento", f"=IFERROR(INDEX(Funcionarios!$G$3:$G${last_func_row},MATCH($C$5,Funcionarios!$A$3:$A${last_func_row},0)),\"\")"),
        ("C. de Custo", f"=IFERROR(INDEX(Funcionarios!$H$3:$H${last_func_row},MATCH($C$5,Funcionarios!$A$3:$A${last_func_row},0)),\"\")"),
        ("Nascimento", f"=IFERROR(INDEX(Funcionarios!$I$3:$I${last_func_row},MATCH($C$5,Funcionarios!$A$3:$A${last_func_row},0)),\"\")"),
        ("Sexo", f"=IFERROR(INDEX(Funcionarios!$J$3:$J${last_func_row},MATCH($C$5,Funcionarios!$A$3:$A${last_func_row},0)),\"\")"),
    ]
    for i, (label, formula) in enumerate(campos):
        r = 7 + i
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=formula)
    ws["B16"] = "Total Liquido Recebido (todas as folhas)"
    ws["C16"] = "=SUMIF(Liquidos!$C:$C,$C$5,Liquidos!$E:$E)"
    ws["B17"] = "Total Horas Extras (valor)"
    ws["C17"] = f"=SUMIF(Horas_Extras!$C:$C,$C$5,Horas_Extras!$H:$H)"
    ws["B18"] = "Faltas (h)"
    ws["C18"] = f"=IFERROR(INDEX(Faltas!$D$3:$D${last_faltas_row},MATCH(INDEX(Funcionarios!$B$3:$B${last_func_row},MATCH($C$5,Funcionarios!$A$3:$A${last_func_row},0)),Faltas!$A$3:$A${last_faltas_row},0)),\"\")"
    ws["B19"] = "Atestados (h)"
    ws["C19"] = f"=IFERROR(INDEX(Atestados!$D$3:$D${last_atestados_row},MATCH(INDEX(Funcionarios!$B$3:$B${last_func_row},MATCH($C$5,Funcionarios!$A$3:$A${last_func_row},0)),Atestados!$A$3:$A${last_atestados_row},0)),\"\")"
    ws["B20"] = "Rescisao (se houver) - Liquido"
    ws["C20"] = f"=IFERROR(INDEX(Rescisoes!$J$3:$J${last_resc_row},MATCH($C$5,Rescisoes!$A$3:$A${last_resc_row},0)),\"Nao rescindido\")"
    for r in range(7, 21):
        ws.cell(row=r, column=2).font = Font(name="Arial", bold=True)
    autosize(ws, {"A": 3, "B": 38, "C": 30, "D": 3})

    # move Capa e Ficha_Funcionario para o inicio
    wb.move_sheet("Capa", offset=-len(wb.sheetnames))
    wb.move_sheet("Ficha_Funcionario", offset=-(len(wb.sheetnames) - 1))

    # limpar metadados
    wb.properties.creator = "github-"
    wb.properties.lastModifiedBy = "github-"
    wb.properties.title = "RH_Folha_Consolidado"
    wb.properties.company = None
    wb.properties.category = None

    out = f"{BASE}/RH_Folha_Consolidado.xlsx"
    wb.save(out)
    print("Salvo:", out)


if __name__ == "__main__":
    build()
